"""Build a transparent annual water-level table from the reservoir workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .common import (
    INTERIM_DIR,
    PROCESSED_DIR,
    RAW_DIR,
    clean_text,
    find_input_file,
    parse_number,
    relative_to_root,
    setup_logging,
)

SOURCE_PATTERN = "Уровни ВДХР*.xlsx"
SOURCE_SHEET = "Данные по участкам"
LOWER_SECTION_ID = "lower_section"
LOWER_SECTION_NAME = "Нижний участок (озерный)"

OUTPUT_COLUMNS = [
    "water_obs_id",
    "site_id",
    "site_name",
    "water_section_id",
    "water_section_name",
    "obs_date",
    "year",
    "water_level_mean_annual_m_abs",
    "water_level_max_annual_m_abs",
    "source_file",
    "source_sheet",
    "source_row",
    "is_missing",
    "missing_reason",
    "qc_flag",
    "qc_note",
]


def write_water_levels_dictionary(path: Path, profile_rows: list[dict[str, object]]) -> None:
    """Write a compact dictionary for the resolved water-level layer."""

    row = profile_rows[0] if profile_rows else {}
    lines = [
        "# Water Levels Dictionary",
        "",
        "This file documents the mechanically extracted water layer used in the current pipeline.",
        "",
        "## Resolved Fields",
        "",
        "- `water_level_mean_annual_m_abs`: annual mean water level for the lower reservoir section, meters absolute.",
        "- `water_level_max_annual_m_abs`: annual maximum water level for the lower reservoir section, meters absolute.",
        "- `obs_date`: intentionally empty because the source workbook provides annual values by year only.",
        f"- `water_section_id`: `{LOWER_SECTION_ID}`.",
        f"- `water_section_name`: `{LOWER_SECTION_NAME}`.",
        "",
        "## Join Logic",
        "",
        "- The lower-section annual series is attached to each shoreline site as shared hydrological context for interval-level joins.",
        "- This preserves pipeline compatibility without inventing site-local observation dates or local water series.",
        "",
    ]
    if row:
        lines.extend(
            [
                "## Extraction Profile",
                "",
                f"- Source workbook: `{row['source_file']}`",
                f"- Source sheet: `{row['source_sheet']}`",
                f"- Source rows scanned: {row['rows_scanned']}",
                f"- Recognized years: {row['recognized_years']}",
                f"- Year range: {row['year_min']}..{row['year_max']}",
                f"- Site rows written per year: {row['site_count']}",
                "",
            ]
        )
    path.write_text("\n".join(lines), encoding="utf-8")


def load_site_rows() -> list[dict[str, str]]:
    """Load processed site rows so the shared water series can join without reworking the pipeline."""

    sites_path = PROCESSED_DIR / "sites.csv"
    if not sites_path.exists():
        raise FileNotFoundError(
            "Expected data/processed/sites.csv before building water levels. Run the site parser first."
        )
    sites_df = pd.read_csv(sites_path, dtype=object)
    return (
        sites_df[["site_id", "site_name"]]
        .dropna(subset=["site_id", "site_name"])
        .drop_duplicates()
        .sort_values(["site_name", "site_id"])
        .to_dict(orient="records")
    )


def find_lower_section_layout(rows: list[tuple[object, ...]]) -> tuple[int, int, int, int] | None:
    """Locate the header rows and lower-section mean/max columns."""

    for header_row_idx, row_values in enumerate(rows[:10], start=1):
        year_col_candidates = [col_idx for col_idx, value in enumerate(row_values, start=1) if clean_text(value) == "Год"]
        if not year_col_candidates:
            continue
        subheader_row_idx = header_row_idx + 1
        if subheader_row_idx > len(rows):
            continue
        subheaders = rows[subheader_row_idx - 1]
        for col_idx, value in enumerate(row_values, start=1):
            if not clean_text(value).startswith("Нижний участок"):
                continue
            mean_label = clean_text(subheaders[col_idx - 1] if col_idx - 1 < len(subheaders) else None)
            max_label = clean_text(subheaders[col_idx] if col_idx < len(subheaders) else None)
            if mean_label == "Н ср" and max_label == "Н макс":
                return header_row_idx, year_col_candidates[0], col_idx, col_idx + 1
    return None


def build_water_levels_raw(output_path: Path | None = None) -> Path:
    """Build `water_levels_raw.csv` and a compact extraction profile."""

    source_path = find_input_file(RAW_DIR, SOURCE_PATTERN)
    output_path = output_path or PROCESSED_DIR / "water_levels_raw.csv"
    profile_path = INTERIM_DIR / "water_levels_profile.json"
    dictionary_path = INTERIM_DIR / "water_levels_manual_dictionary.md"
    logger = setup_logging("build_water_levels_raw")

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    if SOURCE_SHEET not in workbook.sheetnames:
        raise ValueError(f"Expected sheet {SOURCE_SHEET!r} in {source_path.name}.")

    ws = workbook[SOURCE_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    layout = find_lower_section_layout(rows)
    if layout is None:
        raise ValueError(f"Could not locate lower-section mean/max columns in {source_path.name}.")
    header_row_idx, year_col, mean_col, max_col = layout

    site_rows = load_site_rows()
    records: list[dict[str, object]] = []
    source_year_rows = 0
    year_values: list[int] = []

    for row_idx, row_values in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
        row_values = list(row_values)
        year = parse_number(row_values[year_col - 1] if year_col - 1 < len(row_values) else None)
        mean_level = parse_number(row_values[mean_col - 1] if mean_col - 1 < len(row_values) else None)
        max_level = parse_number(row_values[max_col - 1] if max_col - 1 < len(row_values) else None)
        if year is None:
            continue

        year_int = int(year)
        year_values.append(year_int)
        source_year_rows += 1
        is_missing = mean_level is None and max_level is None
        if is_missing:
            missing_reason = "both_levels_missing"
        elif mean_level is None:
            missing_reason = "mean_level_missing"
        elif max_level is None:
            missing_reason = "max_level_missing"
        else:
            missing_reason = None

        qc_flags = [
            "YEAR_ONLY_SOURCE",
            "SHARED_LOWER_SECTION_CONTEXT",
        ]
        qc_notes = [
            "The source workbook provides annual values by year only; no full observation date is available.",
            "The same lower-section annual series is repeated across shoreline sites as shared hydrological context for interval joins.",
        ]
        if missing_reason is not None:
            qc_flags.append("MISSING_WATER_LEVEL_VALUE")
            qc_notes.append("At least one resolved lower-section level value is missing in the source row.")

        for site_row in site_rows:
            records.append(
                {
                    "water_obs_id": f"{site_row['site_id']}_{LOWER_SECTION_ID}_{year_int}_{row_idx}",
                    "site_id": site_row["site_id"],
                    "site_name": site_row["site_name"],
                    "water_section_id": LOWER_SECTION_ID,
                    "water_section_name": LOWER_SECTION_NAME,
                    "obs_date": None,
                    "year": year_int,
                    "water_level_mean_annual_m_abs": mean_level,
                    "water_level_max_annual_m_abs": max_level,
                    "source_file": relative_to_root(source_path),
                    "source_sheet": SOURCE_SHEET,
                    "source_row": row_idx,
                    "is_missing": is_missing,
                    "missing_reason": missing_reason,
                    "qc_flag": ";".join(qc_flags),
                    "qc_note": " ".join(qc_notes),
                }
            )

    result = pd.DataFrame(records, columns=OUTPUT_COLUMNS).sort_values(["site_id", "year", "source_row"])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    result.to_csv(output_path, index=False)

    profile_rows = [
        {
            "source_file": relative_to_root(source_path),
            "source_sheet": SOURCE_SHEET,
            "rows_scanned": max(len(rows) - header_row_idx - 1, 0),
            "recognized_years": source_year_rows,
            "year_min": min(year_values) if year_values else None,
            "year_max": max(year_values) if year_values else None,
            "site_count": len(site_rows),
            "year_column": year_col,
            "mean_column": mean_col,
            "max_column": max_col,
            "water_section_id": LOWER_SECTION_ID,
            "water_section_name": LOWER_SECTION_NAME,
        }
    ]
    profile_path.parent.mkdir(parents=True, exist_ok=True)
    profile_path.write_text(json.dumps(profile_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    write_water_levels_dictionary(dictionary_path, profile_rows)

    logger.info("Built %s with %s rows", relative_to_root(output_path), len(result))
    logger.info("Wrote water-level profile report to %s", relative_to_root(profile_path))
    logger.info("Wrote water-level dictionary to %s", relative_to_root(dictionary_path))
    return output_path


def main() -> None:
    """CLI entrypoint."""

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=PROCESSED_DIR / "water_levels_raw.csv")
    args = parser.parse_args()
    build_water_levels_raw(output_path=args.output)


if __name__ == "__main__":
    main()
